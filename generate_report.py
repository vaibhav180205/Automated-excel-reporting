#!/usr/bin/env python3
"""
Automated Excel Reporting System

This script:
1. Reads sales data from SQLite database
2. Processes data using pandas
3. Generates formatted Excel report with charts
4. Sends report via email

Author: Internship Project
Date: January 2026
"""

import sqlite3
import pandas as pd
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
import configparser
from datetime import datetime
import os
import sys

# ============================================
# SECTION 1: CONFIGURATION LOADING
# ============================================
def load_config():
    """
    Load configuration from config.ini file.
    Returns config object with database and email settings.
    """
    print("[INFO] Loading configuration...")
    config = configparser.ConfigParser()
    
    # Check if config file exists
    if not os.path.exists('config.ini'):
        print("[ERROR] config.ini file not found!")
        print("[INFO] Please create config.ini file with your settings.")
        sys.exit(1)
    
    config.read('config.ini')
    print("[SUCCESS] Configuration loaded successfully.")
    return config

# ============================================
# SECTION 2: DATABASE CONNECTION & QUERYING
# ============================================
def fetch_sales_data(db_path):
    """
    Connect to SQLite database and fetch sales data.
    
    Args:
        db_path: Path to SQLite database file
    
    Returns:
        DataFrame with sales data
    """
    print(f"[INFO] Connecting to database: {db_path}")
    
    try:
        # Create connection to SQLite database
        conn = sqlite3.connect(db_path)
        print("[SUCCESS] Database connection established.")
        
        # SQL Query: Select all sales records
        query = """
        SELECT 
            sale_id,
            sale_date,
            product_name,
            category,
            quantity,
            unit_price,
            total_amount
        FROM sales
        ORDER BY sale_date DESC
        """
        
        print("[INFO] Executing SQL query to fetch sales data...")
        df = pd.read_sql_query(query, conn)
        
        # Close database connection
        conn.close()
        print(f"[SUCCESS] Fetched {len(df)} sales records.")
        
        return df
    
    except Exception as e:
        print(f"[ERROR] Failed to fetch data from database: {e}")
        sys.exit(1)

def get_sales_summary(db_path):
    """
    Get aggregated sales summary using SQL GROUP BY.
    
    Args:
        db_path: Path to SQLite database file
    
    Returns:
        DataFrame with summary statistics
    """
    print("[INFO] Calculating sales summary...")
    
    try:
        conn = sqlite3.connect(db_path)
        
        # SQL Query: Aggregate sales by product
        summary_query = """
        SELECT 
            product_name,
            category,
            COUNT(*) as total_sales,
            SUM(quantity) as total_quantity,
            SUM(total_amount) as total_revenue,
            AVG(total_amount) as avg_sale_amount
        FROM sales
        GROUP BY product_name, category
        ORDER BY total_revenue DESC
        """
        
        df_summary = pd.read_sql_query(summary_query, conn)
        conn.close()
        
        print(f"[SUCCESS] Summary calculated for {len(df_summary)} products.")
        return df_summary
    
    except Exception as e:
        print(f"[ERROR] Failed to calculate summary: {e}")
        sys.exit(1)

# ============================================
# SECTION 3: DATA PROCESSING WITH PANDAS
# ============================================
def process_data(df):
    """
    Clean and process sales data using pandas.
    
    Args:
        df: Raw sales DataFrame
    
    Returns:
        Processed DataFrame
    """
    print("[INFO] Processing data with pandas...")
    
    # Create a copy to avoid modifying original
    df_processed = df.copy()
    
    # Convert sale_date to datetime format
    df_processed['sale_date'] = pd.to_datetime(df_processed['sale_date'])
    
    # Add derived columns
    df_processed['month'] = df_processed['sale_date'].dt.strftime('%Y-%m')
    df_processed['day_of_week'] = df_processed['sale_date'].dt.day_name()
    
    # Round monetary values to 2 decimal places
    df_processed['unit_price'] = df_processed['unit_price'].round(2)
    df_processed['total_amount'] = df_processed['total_amount'].round(2)
    
    # Remove any duplicate records (if any)
    df_processed = df_processed.drop_duplicates()
    
    print(f"[SUCCESS] Data processing complete. Shape: {df_processed.shape}")
    return df_processed

# ============================================
# SECTION 4: EXCEL REPORT GENERATION
# ============================================
def generate_excel_report(df_data, df_summary, output_file):
    """
    Generate formatted Excel report with multiple sheets and charts.
    
    Args:
        df_data: Processed sales data DataFrame
        df_summary: Summary statistics DataFrame
        output_file: Output Excel file path
    """
    print(f"[INFO] Generating Excel report: {output_file}")
    
    try:
        # Create Excel writer object
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            
            # ===== SHEET 1: SUMMARY SHEET =====
            print("[INFO] Creating Summary sheet...")
            
            # Write summary data
            df_summary.to_excel(writer, sheet_name='Summary', index=False, startrow=4)
            
            # Get the workbook and summary sheet
            workbook = writer.book
            summary_sheet = writer.sheets['Summary']
            
            # Add title and metadata
            summary_sheet['A1'] = 'SALES REPORT SUMMARY'
            summary_sheet['A1'].font = Font(size=16, bold=True, color='FFFFFF')
            summary_sheet['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            summary_sheet['A1'].alignment = Alignment(horizontal='center')
            summary_sheet.merge_cells('A1:F1')
            
            summary_sheet['A2'] = f'Generated on: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
            summary_sheet['A2'].font = Font(italic=True)
            
            # Calculate and display totals
            summary_sheet['A3'] = 'Overall Totals:'
            summary_sheet['A3'].font = Font(bold=True)
            summary_sheet['B3'] = f"Total Revenue: ${df_summary['total_revenue'].sum():,.2f}"
            summary_sheet['D3'] = f"Total Quantity: {df_summary['total_quantity'].sum():,}"
            
            # Format header row
            for cell in summary_sheet[5]:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # Set column widths manually
            summary_sheet.column_dimensions['A'].width = 20
            summary_sheet.column_dimensions['B'].width = 15
            summary_sheet.column_dimensions['C'].width = 15
            summary_sheet.column_dimensions['D'].width = 18
            summary_sheet.column_dimensions['E'].width = 18
            summary_sheet.column_dimensions['F'].width = 18
            
            # ===== SHEET 2: RAW DATA SHEET =====
            print("[INFO] Creating Data sheet...")
            
            # Write raw data
            df_data.to_excel(writer, sheet_name='Data', index=False)
            data_sheet = writer.sheets['Data']
            
            # Format header row
            for cell in data_sheet[1]:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # Set column widths for data sheet
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                data_sheet.column_dimensions[col].width = 15
        
        # ===== ADD CHARTS TO SUMMARY SHEET =====
        print("[INFO] Adding charts to report...")
        add_charts_to_report(output_file, len(df_summary))
        
        print(f"[SUCCESS] Excel report generated: {output_file}")
        
    except Exception as e:
        print(f"[ERROR] Failed to generate Excel report: {e}")
        sys.exit(1)

def add_charts_to_report(excel_file, data_rows):
    """
    Add charts to the Excel report.
    
    Args:
        excel_file: Path to Excel file
        data_rows: Number of data rows in summary
    """
    try:
        # Load the workbook
        wb = load_workbook(excel_file)
        summary_sheet = wb['Summary']
        
        # ===== CHART 1: BAR CHART - Revenue by Product =====
        bar_chart = BarChart()
        bar_chart.title = "Revenue by Product"
        bar_chart.x_axis.title = "Product"
        bar_chart.y_axis.title = "Total Revenue ($)"
        bar_chart.style = 10
        
        # Data for chart (columns: product_name and total_revenue)
        products = Reference(summary_sheet, min_col=1, min_row=5, max_row=5+data_rows)
        revenue = Reference(summary_sheet, min_col=5, min_row=4, max_row=5+data_rows)
        
        bar_chart.add_data(revenue, titles_from_data=True)
        bar_chart.set_categories(products)
        bar_chart.height = 10
        bar_chart.width = 20
        
        # Add chart to sheet
        summary_sheet.add_chart(bar_chart, "H5")
        
        # ===== CHART 2: LINE CHART - Quantity Sold by Product =====
        line_chart = LineChart()
        line_chart.title = "Quantity Sold by Product"
        line_chart.x_axis.title = "Product"
        line_chart.y_axis.title = "Total Quantity"
        line_chart.style = 12
        
        # Data for chart
        quantity = Reference(summary_sheet, min_col=4, min_row=4, max_row=5+data_rows)
        
        line_chart.add_data(quantity, titles_from_data=True)
        line_chart.set_categories(products)
        line_chart.height = 10
        line_chart.width = 20
        
        # Add chart to sheet
        summary_sheet.add_chart(line_chart, "H22")
        
        # Save workbook
        wb.save(excel_file)
        print("[SUCCESS] Charts added to report.")
        
    except Exception as e:
        print(f"[ERROR] Failed to add charts: {e}")

# ============================================
# SECTION 5: EMAIL SENDING
# ============================================
def send_email(config, report_file):
    """
    Send Excel report via email using SMTP.
    
    Args:
        config: Configuration object with email settings
        report_file: Path to Excel report file
    """
    print("[INFO] Preparing to send email...")
    
    try:
        # Get email configuration
        sender_email = config['EMAIL']['sender_email']
        sender_password = config['EMAIL']['sender_password']
        receiver_email = config['EMAIL']['receiver_email']
        smtp_server = config['EMAIL']['smtp_server']
        smtp_port = int(config['EMAIL']['smtp_port'])
        
        # Create email message
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = receiver_email
        msg['Subject'] = f"Sales Report - {datetime.now().strftime('%Y-%m-%d')}"
        
        # Email body
        body = f"""
        Hello,
        
        Please find attached the automated sales report generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.
        
        The report includes:
        - Summary sheet with aggregated sales data
        - Detailed data sheet with all transactions
        - Visual charts for quick insights
        
        Best regards,
        Automated Reporting System
        """
        
        msg.attach(MIMEText(body, 'plain'))
        
        # Attach Excel file
        print(f"[INFO] Attaching file: {report_file}")
        with open(report_file, 'rb') as attachment:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header(
                'Content-Disposition',
                f'attachment; filename= {os.path.basename(report_file)}'
            )
            msg.attach(part)
        
        # Connect to SMTP server and send email
        print(f"[INFO] Connecting to SMTP server: {smtp_server}:{smtp_port}")
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()  # Enable encryption
        
        print("[INFO] Logging in to email account...")
        server.login(sender_email, sender_password)
        
        print("[INFO] Sending email...")
        server.send_message(msg)
        server.quit()
        
        print(f"[SUCCESS] Email sent successfully to {receiver_email}")
        
    except Exception as e:
        print(f"[ERROR] Failed to send email: {e}")
        print("[INFO] Report was still generated successfully as Excel file.")

# ============================================
# SECTION 6: MAIN EXECUTION
# ============================================
def main():
    """
    Main function to orchestrate the entire reporting process.
    """
    print("="*60)
    print("AUTOMATED EXCEL REPORTING SYSTEM")
    print("="*60)
    print(f"Started at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("="*60)
    
    try:
        # Step 1: Load configuration
        config = load_config()
        
        # Step 2: Get database path from config
        db_path = config['DATABASE']['db_path']
        
        # Step 3: Fetch sales data from database
        df_raw = fetch_sales_data(db_path)
        
        # Step 4: Get sales summary with aggregations
        df_summary = get_sales_summary(db_path)
        
        # Step 5: Process data with pandas
        df_processed = process_data(df_raw)
        
        # Step 6: Generate output filename
        output_file = config['REPORT']['output_path']
        # Add timestamp to filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = output_file.replace('.xlsx', f'_{timestamp}.xlsx')
        
        # Step 7: Generate Excel report
        generate_excel_report(df_processed, df_summary, output_file)
        
        # Step 8: Send email (if configured)
        if config['EMAIL'].getboolean('send_email', fallback=True):
            send_email(config, output_file)
        else:
            print("[INFO] Email sending is disabled in config.")
        
        print("="*60)
        print("REPORT GENERATION COMPLETED SUCCESSFULLY!")
        print(f"Report saved at: {output_file}")
        print("="*60)
        
    except Exception as e:
        print("="*60)
        print(f"[FATAL ERROR] Program failed: {e}")
        print("="*60)
        sys.exit(1)

if __name__ == "__main__":
    main()
